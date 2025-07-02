from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# CORS setup to allow only your Shopify domain
CORS(app, resources={r"/api/*": {
    "origins": "https://zhen-first-store-com.myshopify.com",
    "allow_headers": ["Content-Type"],
    "methods": ["POST", "OPTIONS"]
}})

EXCEL_FILE = 'subscribers.xlsx'

# Initialize Excel file if it doesn't exist
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email"])
        wb.save(EXCEL_FILE)

@app.route('/api/register', methods=['POST', 'OPTIONS'])
def register():
    # Handle preflight CORS request
    if request.method == 'OPTIONS':
        response = app.make_response('')
        response.headers['Access-Control-Allow-Origin'] = 'https://zhen-first-store-com.myshopify.com'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    try:
        data = request.get_json(force=True)
        name = data.get('name')
        email = data.get('email')

        if not name or not email:
            return jsonify({"error": "Name and email required"}), 400

        if '@' not in email or '.' not in email:
            return jsonify({"error": "Invalid email format"}), 400

        # Append to Excel file
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Email"])

        ws.append([name, email])
        wb.save(EXCEL_FILE)

        return jsonify({"message": "Registered successfully"}), 200

    except Exception as e:
        print("Error:", e)
        return jsonify({"error": "Server error"}), 500

if __name__ == '__main__':
    init_excel()
    app.run(port=5000)
